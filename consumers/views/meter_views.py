from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import ListView
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.conf import settings
from ..decorators import (
    get_client_ip, get_user_agent, is_admin_user, is_superuser_only,
    consumer_edit_permission_required, disconnect_permission_required,
    user_management_permission_required, system_settings_permission_required,
    billing_permission_required, reports_permission_required, view_only_for_admin,
    rate_limit_login, role_required
)
from django.db.models import Q, Max, Count, Sum, OuterRef, Subquery, Value, F
from django.db.models.functions import Concat, TruncMonth
from django.db import models
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone
from django.utils.decorators import method_decorator
from django.urls import reverse
from django.core.paginator import Paginator
from django.contrib.auth.models import User
from datetime import datetime, timedelta, date
try:
    from dateutil.relativedelta import relativedelta
except Exception:
    # Fallback: approximate relativedelta by using a timedelta of ~30 days per month
    # This keeps existing subtraction usages like relativedelta(months=5) working
    from datetime import timedelta as _td
    def relativedelta(months=0, **kwargs):
        return _td(days=30 * int(months))
from decimal import Decimal, InvalidOperation
import uuid
import json
import csv
import base64
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Cloudinary import with error handling (optional dependency)
try:
    from cloudinary import uploader as cloudinary_uploader  # type: ignore
    CLOUDINARY_AVAILABLE = True
except ImportError:
    cloudinary_uploader = None
    CLOUDINARY_AVAILABLE = False

from ..models import (
    Consumer, Barangay, Purok, MeterReading, Bill, SystemSetting, Payment,
    StaffProfile, UserLoginEvent, MeterBrand, PasswordResetToken, UserActivity,
    SystemSettingChangeLog, Notification
)
from ..forms import ConsumerForm


# Helper function to authenticate API requests using session token
def authenticate_api_request(request):
    """
    Authenticate API request using session token from Authorization header or request body.
    Returns the user if authenticated, None otherwise.
    """
    from django.contrib.sessions.models import Session

    token = None

    # Try Authorization header first (Bearer token)
    auth_header = request.META.get('HTTP_AUTHORIZATION', '')
    if auth_header.startswith('Bearer '):
        token = auth_header[7:]

    # Try request body if no header token
    if not token and request.body:
        try:
            data = json.loads(request.body.decode('utf-8'))
            token = data.get('token')
        except (json.JSONDecodeError, UnicodeDecodeError):
            pass

    if not token:
        return None

    try:
        # Find the session by key
        session = Session.objects.get(session_key=token)

        # Check if session is expired
        if session.expire_date < timezone.now():
            return None

        # Get user from session data
        session_data = session.get_decoded()
        user_id = session_data.get('_auth_user_id')

        if user_id:
            user = User.objects.get(id=user_id)
            return user
    except (Session.DoesNotExist, User.DoesNotExist):
        pass

    return None


# Helper function to get previous confirmed reading
def get_previous_reading(consumer):
    """Get the most recent confirmed meter reading for a consumer."""
    latest_reading = MeterReading.objects.filter(
        consumer=consumer,
        is_confirmed=True
    ).order_by('-reading_date', '-created_at').first()

    return latest_reading.reading_value if latest_reading else 0


# Helper function to calculate water bill
def calculate_water_bill(consumer, consumption):
    """
    Calculate water bill using TIERED rate structure from System Settings.

    Returns: (average_rate, total_amount, breakdown)
    - average_rate: Effective rate per cubic meter
    - total_amount: Total bill amount
    - breakdown: Dict with tier-by-tier calculation details
    """
    from ..utils import calculate_tiered_water_bill

    # Use tiered calculation from utils
    total_amount, average_rate, breakdown = calculate_tiered_water_bill(
        consumption=consumption,
        usage_type=consumer.usage_type
    )

    return float(average_rate), float(total_amount), breakdown




def meter_reading_overview(request):
    """
    Meter Reading Overview with Enhanced Statistics and Progress Tracking.

    Features:
    - Barangay-level summary with progress bars
    - Overall completion percentage
    - Excel export for summary report
    - Updated vs pending consumer counts
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from django.http import HttpResponse

    import calendar

    today = date.today()
    export_excel = request.GET.get('export', '')

    # Month filter: accept ?month=2026-01 format, default to current month
    month_param = request.GET.get('month', '')
    if month_param:
        try:
            year, mon = month_param.split('-')
            current_month = date(int(year), int(mon), 1)
        except (ValueError, TypeError):
            current_month = today.replace(day=1)
    else:
        current_month = today.replace(day=1)

    # Calculate the end of the selected month
    last_day = calendar.monthrange(current_month.year, current_month.month)[1]
    month_end = date(current_month.year, current_month.month, last_day)

    # Get all barangays, annotate total consumers, and ORDER BY name for alphabetical sorting
    barangays = Barangay.objects.annotate(
        total_consumers=Count('consumer', distinct=True)
    ).order_by('name')

    barangay_data = []
    for b in barangays:
        # Only count active (connected) consumers
        consumer_ids = list(Consumer.objects.filter(barangay=b, status='active').values_list('id', flat=True))
        total_consumers_count = len(consumer_ids)

        if not consumer_ids:
            ready = not_updated = updated = 0
            completion_percentage = 0
        else:
            # Count unconfirmed readings for the selected month
            ready = MeterReading.objects.filter(
                consumer_id__in=consumer_ids,
                is_confirmed=False,
                reading_date__gte=current_month,
                reading_date__lte=month_end
            ).values('consumer_id').distinct().count()

            # Count consumers who have at least one reading in the selected month
            updated_consumers = MeterReading.objects.filter(
                consumer_id__in=consumer_ids,
                reading_date__gte=current_month,
                reading_date__lte=month_end
            ).values_list('consumer_id', flat=True).distinct()
            updated = len(set(updated_consumers))
            not_updated = total_consumers_count - updated

            # Calculate completion percentage (based on updated readings)
            completion_percentage = (updated / total_consumers_count * 100) if total_consumers_count > 0 else 0

        barangay_data.append({
            'barangay': b,
            'ready_to_confirm': ready,
            'not_yet_updated': not_updated,
            'updated_count': updated,
            'total_consumers': total_consumers_count,
            'completion_percentage': round(completion_percentage, 1),
        })

    # Calculate summary statistics for the overview page
    total_barangays = len(barangay_data)
    total_consumers_sum = sum(item['total_consumers'] for item in barangay_data)
    total_ready_sum = sum(item['ready_to_confirm'] for item in barangay_data)
    total_pending_sum = sum(item['not_yet_updated'] for item in barangay_data)
    total_updated_sum = sum(item['updated_count'] for item in barangay_data)

    # Calculate overall completion percentage
    overall_completion_percentage = (total_updated_sum / total_consumers_sum * 100) if total_consumers_sum > 0 else 0

    # Print Report Export (HTML with 2 logos)
    if export_excel == 'print':
        context = {
            'barangay_data': barangay_data,
            'current_month': current_month,
            'total_barangays': total_barangays,
            'total_consumers_sum': total_consumers_sum,
            'total_updated_sum': total_updated_sum,
            'total_pending_sum': total_pending_sum,
            'overall_completion_percentage': round(overall_completion_percentage, 1),
            'generated_date': timezone.now(),
        }
        return render(request, 'consumers/meter_reading_overview_print.html', context)

    # Excel Export
    if export_excel == 'excel':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reading Overview"

        # Header styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)

        # Title
        ws.merge_cells('A1:E1')
        title_cell = ws['A1']
        title_cell.value = f"Meter Reading Overview - {current_month.strftime('%B %Y')}"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')

        # Headers
        headers = ['Barangay', 'Total Consumers', 'Readings This Month', 'Not Updated', 'Progress %']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Data rows
        for idx, item in enumerate(barangay_data, 4):
            ws.cell(row=idx, column=1, value=item['barangay'].name)
            ws.cell(row=idx, column=2, value=item['total_consumers'])
            ws.cell(row=idx, column=3, value=item['updated_count'])
            ws.cell(row=idx, column=4, value=item['not_yet_updated'])
            ws.cell(row=idx, column=5, value=f"{item['completion_percentage']}%")

        # Summary row
        summary_row = len(barangay_data) + 5
        ws.cell(row=summary_row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=summary_row, column=2, value=total_consumers_sum).font = Font(bold=True)
        ws.cell(row=summary_row, column=3, value=total_updated_sum).font = Font(bold=True)
        ws.cell(row=summary_row, column=4, value=total_pending_sum).font = Font(bold=True)
        ws.cell(row=summary_row, column=5, value=f"{round(overall_completion_percentage, 1)}%").font = Font(bold=True)

        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        for col in ['B', 'C', 'D', 'E']:
            ws.column_dimensions[col].width = 18

        # Create response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=meter_reading_overview_{today.strftime("%Y%m%d")}.xlsx'
        wb.save(response)
        return response

    # Format month param for template links
    month_filter_value = current_month.strftime('%Y-%m')

    context = {
        'barangay_data': barangay_data,
        'current_month': current_month,
        'month_filter_value': month_filter_value,
        'total_barangays': total_barangays,
        'total_consumers_sum': total_consumers_sum,
        'total_ready_sum': total_ready_sum,
        'total_pending_sum': total_pending_sum,
        'total_updated_sum': total_updated_sum,
        'overall_completion_percentage': round(overall_completion_percentage, 1),
    }

    return render(request, 'consumers/meter_reading_overview.html', context)


@login_required
def barangay_meter_readings(request, barangay_id):
    import calendar

    barangay = get_object_or_404(Barangay, id=barangay_id)
    today = date.today()

    # Month filter: accept ?month=2026-01 format
    month_param = request.GET.get('month', '')
    filter_month = None
    month_start = None
    month_end = None
    if month_param:
        try:
            year, mon = month_param.split('-')
            month_start = date(int(year), int(mon), 1)
            last_day = calendar.monthrange(month_start.year, month_start.month)[1]
            month_end = date(month_start.year, month_start.month, last_day)
            filter_month = month_start
        except (ValueError, TypeError):
            pass

    # Get all active consumers in this barangay
    consumers = Consumer.objects.filter(barangay=barangay, status='active').select_related('barangay').order_by('id')

    readings_with_data = []
    for consumer in consumers:
        if filter_month:
            # Get latest reading for this consumer within the filtered month
            latest_reading = MeterReading.objects.filter(
                consumer=consumer,
                reading_date__gte=month_start,
                reading_date__lte=month_end
            ).order_by('-reading_date', '-created_at').first()
        else:
            # Get latest reading for this consumer (unconfirmed takes priority for display)
            latest_reading = MeterReading.objects.filter(consumer=consumer).order_by('-reading_date', '-created_at').first()

        if latest_reading:
            # Find previous confirmed reading
            prev = MeterReading.objects.filter(
                consumer=consumer,
                is_confirmed=True,
                reading_date__lt=latest_reading.reading_date
            ).order_by('-reading_date').first()

            # Calculate consumption
            if prev:
                consumption = latest_reading.reading_value - prev.reading_value
            else:
                # First reading - use consumer's first_reading as baseline
                baseline = consumer.first_reading if consumer.first_reading else 0
                consumption = latest_reading.reading_value - baseline
        else:
            latest_reading = None
            prev = None
            consumption = None

        readings_with_data.append({
            'consumer': consumer,
            'reading': latest_reading,
            'prev_reading': prev,
            'consumption': consumption,
            'display_id': get_consumer_display_id(consumer),
        })

    # Calculate counts for summary statistics
    pending_count = sum(1 for item in readings_with_data if item['reading'] and not item['reading'].is_confirmed)
    confirmed_count = sum(1 for item in readings_with_data if item['reading'] and item['reading'].is_confirmed)
    no_reading_count = sum(1 for item in readings_with_data if not item['reading'])

    return render(request, 'consumers/barangay_meter_readings.html', {
        'barangay': barangay,
        'readings': readings_with_data,
        'today': today,
        'pending_count': pending_count,
        'confirmed_count': confirmed_count,
        'no_reading_count': no_reading_count,
        'filter_month': filter_month,
        'month_filter_value': month_param,
    })



# ───────────────────────────────────────
# PRINT: Barangay Meter Readings Report
# ───────────────────────────────────────
@login_required
def barangay_meter_readings_print(request, barangay_id):
    """Printable barangay meter readings report with 2 logos"""
    from django.utils import timezone

    barangay = get_object_or_404(Barangay, id=barangay_id)
    today = date.today()
    current_month = today.replace(day=1)

    # Get all active consumers in this barangay
    consumers = Consumer.objects.filter(barangay=barangay, status='active').select_related('barangay').order_by('id')

    readings_with_data = []
    for consumer in consumers:
        # Get latest reading for this consumer
        latest_reading = MeterReading.objects.filter(consumer=consumer).order_by('-reading_date', '-created_at').first()

        if latest_reading:
            # Find previous confirmed reading
            prev = MeterReading.objects.filter(
                consumer=consumer,
                is_confirmed=True,
                reading_date__lt=latest_reading.reading_date
            ).order_by('-reading_date').first()

            # Calculate consumption
            if prev:
                consumption = latest_reading.reading_value - prev.reading_value
            else:
                # First reading - use consumer's first_reading as baseline
                baseline = consumer.first_reading if consumer.first_reading else 0
                consumption = latest_reading.reading_value - baseline
        else:
            latest_reading = None
            prev = None
            consumption = None

        readings_with_data.append({
            'consumer': consumer,
            'reading': latest_reading,
            'prev_reading': prev,
            'consumption': consumption,
            'display_id': get_consumer_display_id(consumer),
        })

    # Calculate summary statistics
    total_consumers = len(readings_with_data)
    with_readings = sum(1 for item in readings_with_data if item['reading'])
    no_readings = sum(1 for item in readings_with_data if not item['reading'])

    context = {
        'barangay': barangay,
        'readings': readings_with_data,
        'total_consumers': total_consumers,
        'with_readings': with_readings,
        'no_readings': no_readings,
        'current_month': current_month,
        'generated_date': timezone.now(),
    }

    return render(request, 'consumers/barangay_meter_readings_print.html', context)



# ───────────────────────────────────────
# NEW: Confirm All Readings in Barangay
# ───────────────────────────────────────
@login_required
def confirm_all_readings(request, barangay_id):
    if request.method != "POST":
        return redirect('consumers:barangay_meter_readings', barangay_id=barangay_id)

    barangay = get_object_or_404(Barangay, id=barangay_id)
    readings_to_confirm = MeterReading.objects.filter(
        consumer__barangay=barangay,
        is_confirmed=False
    ).select_related('consumer')

    success_count = 0
    for reading in readings_to_confirm:
        try:
            # Find previous confirmed reading
            prev = MeterReading.objects.filter(
                consumer=reading.consumer,
                is_confirmed=True,
                reading_date__lt=reading.reading_date
            ).order_by('-reading_date').first()

            # Calculate consumption
            if prev:
                # Has previous reading - validate current >= previous
                if reading.reading_value < prev.reading_value:
                    continue
                cons = reading.reading_value - prev.reading_value
            else:
                # First reading for this consumer - use consumer's first_reading as baseline
                baseline = reading.consumer.first_reading if reading.consumer.first_reading else 0
                if reading.reading_value < baseline:
                    continue
                cons = reading.reading_value - baseline

            # Get system settings and calculate using TIERED RATES
            from ..utils import calculate_tiered_water_bill

            setting = SystemSetting.objects.first()
            if setting:
                billing_day = setting.billing_day_of_month
                due_day = setting.due_day_of_month
            else:
                billing_day = 1
                due_day = 20

            # Calculate bill using TIERED RATES
            total, average_rate, breakdown = calculate_tiered_water_bill(
                consumption=cons,
                usage_type=reading.consumer.usage_type,
                settings=setting
            )

            bill = Bill.objects.create(
                consumer=reading.consumer,
                previous_reading=prev,
                current_reading=reading,
                billing_period=reading.reading_date.replace(day=billing_day),
                due_date=reading.reading_date.replace(day=due_day),
                consumption=cons,
                # Store ACTUAL tier breakdown
                tier1_consumption=breakdown['tier1_units'],
                tier1_amount=breakdown['tier1_amount'],
                tier2_consumption=breakdown['tier2_units'],
                tier2_rate=breakdown['tier2_rate'],
                tier2_amount=breakdown['tier2_amount'],
                tier3_consumption=breakdown['tier3_units'],
                tier3_rate=breakdown['tier3_rate'],
                tier3_amount=breakdown['tier3_amount'],
                tier4_consumption=breakdown['tier4_units'],
                tier4_rate=breakdown['tier4_rate'],
                tier4_amount=breakdown['tier4_amount'],
                tier5_consumption=breakdown['tier5_units'],
                tier5_rate=breakdown['tier5_rate'],
                tier5_amount=breakdown['tier5_amount'],
                rate_per_cubic=average_rate,
                fixed_charge=Decimal('0.00'),
                total_amount=total,
                status='Pending'
            )
            
            # Send SMS Bill Alert
            from ..utils import send_bill_sms
            send_bill_sms(bill)
            
            reading.is_confirmed = True
            reading.save()
            success_count += 1
        except Exception as e:
            import logging
            logging.error(f"Error confirming reading {reading.id}: {str(e)}")
            continue

    if success_count > 0:
        messages.success(request, f"✅ {success_count} readings confirmed and bills generated.")
    return redirect('consumers:barangay_meter_readings', barangay_id=barangay_id)



@login_required
def confirm_all_readings_global(request):
    """
    Confirm ALL unconfirmed readings across ALL barangays.
    """
    if request.method != "POST":
        return redirect('consumers:meter_readings')

    # Get all unconfirmed readings
    readings_to_confirm = MeterReading.objects.filter(
        is_confirmed=False
    ).select_related('consumer')

    success_count = 0
    error_count = 0

    for reading in readings_to_confirm:
        try:
            consumer = reading.consumer

            # Find previous confirmed reading
            prev = MeterReading.objects.filter(
                consumer=consumer,
                is_confirmed=True,
                reading_date__lt=reading.reading_date
            ).order_by('-reading_date').first()

            # Calculate consumption
            if prev:
                if reading.reading_value < prev.reading_value:
                    continue
                cons = reading.reading_value - prev.reading_value
            else:
                # First reading for this consumer - use consumer's first_reading as baseline
                baseline = consumer.first_reading if consumer.first_reading else 0
                if reading.reading_value < baseline:
                    continue
                cons = reading.reading_value - baseline

            # Get system settings and calculate using TIERED RATES
            from ..utils import calculate_tiered_water_bill

            setting = SystemSetting.objects.first()
            if setting:
                billing_day = setting.billing_day_of_month
                due_day = setting.due_day_of_month
            else:
                billing_day = 1
                due_day = 20

            # Calculate bill using TIERED RATES
            total, average_rate, breakdown = calculate_tiered_water_bill(
                consumption=cons,
                usage_type=consumer.usage_type,
                settings=setting
            )

            bill = Bill.objects.create(
                consumer=consumer,
                previous_reading=prev,
                current_reading=reading,
                billing_period=reading.reading_date.replace(day=billing_day),
                due_date=reading.reading_date.replace(day=due_day),
                consumption=cons,
                # Store ACTUAL tier breakdown
                tier1_consumption=breakdown['tier1_units'],
                tier1_amount=breakdown['tier1_amount'],
                tier2_consumption=breakdown['tier2_units'],
                tier2_rate=breakdown['tier2_rate'],
                tier2_amount=breakdown['tier2_amount'],
                tier3_consumption=breakdown['tier3_units'],
                tier3_rate=breakdown['tier3_rate'],
                tier3_amount=breakdown['tier3_amount'],
                tier4_consumption=breakdown['tier4_units'],
                tier4_rate=breakdown['tier4_rate'],
                tier4_amount=breakdown['tier4_amount'],
                tier5_consumption=breakdown['tier5_units'],
                tier5_rate=breakdown['tier5_rate'],
                tier5_amount=breakdown['tier5_amount'],
                rate_per_cubic=average_rate,
                fixed_charge=Decimal('0.00'),
                total_amount=total,
                status='Pending'
            )
            
            # Send SMS Bill Alert
            from ..utils import send_bill_sms
            send_bill_sms(bill)
            
            reading.is_confirmed = True
            reading.save()
            success_count += 1
        except Exception as e:
            import logging
            logging.error(f"Error confirming reading {reading.id}: {str(e)}")
            error_count += 1
            continue

    if success_count > 0:
        messages.success(request, f"✅ {success_count} readings confirmed and bills generated across all barangays.")
    else:
        messages.info(request, "No unconfirmed readings found.")

    if error_count > 0:
        messages.warning(request, f"⚠️ {error_count} readings could not be processed.")

    return redirect('consumers:meter_readings')



# ───────────────────────────────────────
# NEW: Export to Excel
# ───────────────────────────────────────
@login_required
def export_barangay_readings(request, barangay_id):
    barangay = get_object_or_404(Barangay, id=barangay_id)
    current_month = date.today().replace(day=1)

    # Get latest reading per consumer in this barangay
    consumer_ids = Consumer.objects.filter(barangay=barangay).values_list('id', flat=True)
    if not consumer_ids:
        readings = MeterReading.objects.none()
    else:
        latest_date_subq = MeterReading.objects.filter(
            consumer=OuterRef('pk')
        ).order_by().values('consumer').annotate(
            max_date=Max('reading_date')
        ).values('max_date')[:1]

        readings = MeterReading.objects.select_related('consumer').filter(
            consumer__barangay=barangay,
            reading_date=Subquery(latest_date_subq)
        ).order_by('consumer__id')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{barangay.name} Readings"

    headers = [
        'ID Number',           # ← Changed from "Account ID"
        'Consumer Name',
        'Current',
        'Previous',
        'Consumption (m³)',
        'Date',
        'Status'
    ]
    ws.append(headers)

    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col in range(1, len(headers) + 1):
        ws.cell(1, col).fill = header_fill
        ws.cell(1, col).font = header_font

    for r in readings:
        prev = MeterReading.objects.filter(
            consumer=r.consumer,
            is_confirmed=True,
            reading_date__lt=r.reading_date
        ).order_by('-reading_date').first()
        cons = (r.reading_value - prev.reading_value) if prev else '—'
        display_id = get_consumer_display_id(r.consumer)

        ws.append([
            display_id,
            f"{r.consumer.first_name} {r.consumer.last_name}",
            r.reading_value,
            prev.reading_value if prev else '—',
            cons,
            r.reading_date.strftime('%Y-%m-%d'),
            'Confirmed' if r.is_confirmed else 'Pending'
        ])

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column].width = adjusted_width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=Readings_{barangay.name}_{current_month.strftime("%Y-%m")}.xlsx'
    wb.save(response)
    return response


# ============================================================================
# CONFIRM READING VIEW - ADMIN MANUAL CONFIRMATION
# ============================================================================
# This is the CRITICAL step where admin manually confirms a meter reading.
# Upon confirmation:
# 1. Consumption is calculated (current - previous reading)
# 2. Bill is IMMEDIATELY generated with status='Pending'
# 3. Bill appears in Inquire/Payment page for payment processing
#
# TESTING FLOW:
# Step 1: Mobile app submits reading (is_confirmed=False)
# Step 2: Admin clicks "Confirm" button HERE → Bill created instantly
# Step 3: Admin goes to Inquire page → Pays the bill
#
# NOTE: billing_day_of_month and due_day_of_month from SystemSettings
# only set the DATES on the bill, they do NOT delay bill creation.
# Bill is created IMMEDIATELY when you click Confirm.
# ============================================================================
@login_required
def confirm_reading(request, reading_id):
    """
    Confirm a meter reading and generate a bill for the consumer.
    Supports both regular requests and AJAX requests.
    """
    # Helper to check if AJAX request
    def is_ajax_request():
        return request.headers.get('X-Requested-With') == 'XMLHttpRequest' or \
               'application/json' in request.headers.get('Content-Type', '') or \
               'application/json' in request.headers.get('Accept', '')

    # Helper to return error (AJAX or redirect)
    def return_error(message, barangay_id):
        if is_ajax_request():
            return JsonResponse({'status': 'error', 'message': message}, status=400)
        messages.error(request, message)
        return redirect('consumers:barangay_meter_readings', barangay_id=barangay_id)

    try:
        current = get_object_or_404(MeterReading, id=reading_id)
        consumer = current.consumer
        barangay_id = consumer.barangay.id

        # Check if already confirmed
        if current.is_confirmed:
            return return_error("This reading is already confirmed and billed.", barangay_id)

        # Validate date
        if current.reading_date > date.today():
            return return_error("Reading date cannot be in the future.", barangay_id)

        # Find previous confirmed reading
        previous = MeterReading.objects.filter(
            consumer=consumer,
            is_confirmed=True,
            reading_date__lt=current.reading_date
        ).order_by('-reading_date').first()

        # Calculate consumption
        if previous:
            if current.reading_value < previous.reading_value:
                return return_error(f"Current reading ({current.reading_value}) cannot be less than previous ({previous.reading_value}).", barangay_id)
            consumption = current.reading_value - previous.reading_value
        else:
            baseline = consumer.first_reading if consumer.first_reading else 0
            if current.reading_value < baseline:
                return return_error(f"Current reading ({current.reading_value}) cannot be less than initial ({baseline}).", barangay_id)
            consumption = current.reading_value - baseline

        # Generate bill
        from ..utils import calculate_tiered_water_bill
        setting = SystemSetting.objects.first()

        billing_day = setting.billing_day_of_month if setting else 1
        due_day = setting.due_day_of_month if setting else 20

        total_amount, average_rate, breakdown = calculate_tiered_water_bill(
            consumption=consumption,
            usage_type=consumer.usage_type,
            settings=setting
        )

        bill = Bill.objects.create(
            consumer=consumer,
            previous_reading=previous,
            current_reading=current,
            billing_period=current.reading_date.replace(day=billing_day),
            due_date=current.reading_date.replace(day=due_day),
            consumption=consumption,
            tier1_consumption=breakdown['tier1_units'],
            tier1_amount=breakdown['tier1_amount'],
            tier2_consumption=breakdown['tier2_units'],
            tier2_rate=breakdown['tier2_rate'],
            tier2_amount=breakdown['tier2_amount'],
            tier3_consumption=breakdown['tier3_units'],
            tier3_rate=breakdown['tier3_rate'],
            tier3_amount=breakdown['tier3_amount'],
            tier4_consumption=breakdown['tier4_units'],
            tier4_rate=breakdown['tier4_rate'],
            tier4_amount=breakdown['tier4_amount'],
            tier5_consumption=breakdown['tier5_units'],
            tier5_rate=breakdown['tier5_rate'],
            tier5_amount=breakdown['tier5_amount'],
            rate_per_cubic=average_rate,
            fixed_charge=Decimal('0.00'),
            total_amount=total_amount,
            status='Pending'
        )

        # Send SMS Bill Alert
        from ..utils import send_bill_sms
        send_bill_sms(bill)

        # Mark reading as confirmed
        current.is_confirmed = True
        current.confirmed_by = request.user
        current.confirmed_at = timezone.now()
        current.save()

        success_message = f"Bill generated for {get_consumer_display_id(consumer)}!"

        if is_ajax_request():
            return JsonResponse({'status': 'success', 'message': success_message})

        messages.success(request, f"✅ {success_message}")
        return redirect('consumers:barangay_meter_readings', barangay_id=barangay_id)

    except Exception as e:
        error_message = f"Failed to generate bill: {str(e)}"
        if is_ajax_request():
            return JsonResponse({'status': 'error', 'message': error_message}, status=400)
        messages.error(request, error_message)
        return redirect('consumers:meter_reading_overview')



@login_required
def reject_reading(request, reading_id):
    """
    Reject a meter reading submitted with proof photo.
    Only accessible via POST with a rejection reason.
    """
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Method not allowed'}, status=405)

    reading = get_object_or_404(MeterReading, id=reading_id)

    # Check if already processed
    if reading.is_confirmed:
        return JsonResponse({'status': 'error', 'message': 'This reading is already confirmed'}, status=400)
    if reading.is_rejected:
        return JsonResponse({'status': 'error', 'message': 'This reading is already rejected'}, status=400)

    # Get rejection reason from request body
    try:
        data = json.loads(request.body)
        reason = data.get('reason', '').strip()
    except (json.JSONDecodeError, AttributeError):
        reason = request.POST.get('reason', '').strip()

    if not reason:
        return JsonResponse({'status': 'error', 'message': 'Rejection reason is required'}, status=400)

    # Mark as rejected
    reading.is_rejected = True
    reading.rejected_by = request.user
    reading.rejected_at = timezone.now()
    reading.rejection_reason = reason
    reading.save()

    # Create notification for the field staff who submitted
    if reading.submitted_by:
        from django.urls import reverse
        Notification.objects.create(
            user=reading.submitted_by,
            notification_type='reading_rejected',
            title='Reading Rejected',
            message=f"Your reading for {reading.consumer.first_name} {reading.consumer.last_name} was rejected: {reason}",
            redirect_url=reverse('consumers:consumer_bill', args=[reading.consumer.id])
        )

    return JsonResponse({
        'status': 'success',
        'message': f'Reading rejected for {reading.consumer.first_name} {reading.consumer.last_name}'
    })



@login_required
def meter_readings_print(request):
    """Printable meter readings report with receipt-style header"""
    from datetime import datetime

    # Get filter parameters
    search_query = request.GET.get('search', '').strip()
    selected_barangay = request.GET.get('barangay', '')
    selected_status = request.GET.get('status', '')
    from_date = request.GET.get('from_date', '')
    to_date = request.GET.get('to_date', '')

    # Build queryset with same filters as the view
    readings_queryset = MeterReading.objects.select_related(
        'consumer', 'consumer__barangay', 'submitted_by'
    ).order_by('-reading_date', '-created_at')

    if search_query:
        readings_queryset = readings_queryset.filter(
            Q(consumer__first_name__icontains=search_query) |
            Q(consumer__last_name__icontains=search_query) |
            Q(consumer__id_number__icontains=search_query)
        )

    if selected_barangay:
        readings_queryset = readings_queryset.filter(consumer__barangay_id=selected_barangay)

    if selected_status:
        if selected_status == 'confirmed':
            readings_queryset = readings_queryset.filter(is_confirmed=True)
        elif selected_status == 'pending':
            readings_queryset = readings_queryset.filter(is_confirmed=False, is_rejected=False)

    if from_date:
        readings_queryset = readings_queryset.filter(reading_date__gte=from_date)

    if to_date:
        readings_queryset = readings_queryset.filter(reading_date__lte=to_date)

    # Limit to 500 records for print
    readings_queryset = readings_queryset[:500]

    # Prepare readings with consumption data
    readings_with_data = []
    for reading in readings_queryset:
        prev_reading = MeterReading.objects.filter(
            consumer=reading.consumer,
            is_confirmed=True,
            reading_date__lt=reading.reading_date
        ).order_by('-reading_date').first()

        if prev_reading:
            consumption = reading.reading_value - prev_reading.reading_value
            prev_value = prev_reading.reading_value
        elif reading.consumer.first_reading:
            consumption = reading.reading_value - reading.consumer.first_reading
            prev_value = reading.consumer.first_reading
        else:
            consumption = reading.reading_value
            prev_value = 0

        readings_with_data.append({
            'reading': reading,
            'prev_value': prev_value,
            'consumption': consumption if reading.is_confirmed else (consumption if consumption >= 0 else 0),
        })

    # Build filter display
    filter_display_parts = []
    if selected_barangay:
        try:
            barangay = Barangay.objects.get(id=selected_barangay)
            filter_display_parts.append(f"Barangay: {barangay.name}")
        except:
            pass
    if selected_status:
        filter_display_parts.append(f"Status: {selected_status.title()}")
    if from_date and to_date:
        filter_display_parts.append(f"Period: {from_date} to {to_date}")
    elif from_date:
        filter_display_parts.append(f"From: {from_date}")
    elif to_date:
        filter_display_parts.append(f"To: {to_date}")

    filter_display = " | ".join(filter_display_parts) if filter_display_parts else "All Records"

    # Calculate statistics
    total_count = len(readings_with_data)
    confirmed_count = sum(1 for item in readings_with_data if item['reading'].is_confirmed)
    pending_count = sum(1 for item in readings_with_data if not item['reading'].is_confirmed and not item['reading'].is_rejected)

    context = {
        'readings_with_data': readings_with_data,
        'filter_display': filter_display,
        'total_count': total_count,
        'confirmed_count': confirmed_count,
        'pending_count': pending_count,
        'generated_date': timezone.now()
    }

    return render(request, 'consumers/meter_readings_print.html', context)



class MeterReadingListView(LoginRequiredMixin, ListView):
    """
    Unified meter readings management view with tabbed interface using CBV.
    - All Readings tab: Shows all readings with filters
    - Pending Review tab: Shows only unconfirmed readings
    """
    model = MeterReading
    template_name = 'consumers/meter_readings.html'
    context_object_name = 'readings'

    def get_queryset(self):
        # We don't use standard pagination here because of the complex array-building logic
        # Instead, we'll build the reading list in get_context_data to maintain the existing format
        return MeterReading.objects.none()
        
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        today = date.today()

        # Get all barangays for filter
        barangays = Barangay.objects.all().order_by('name')

        # Subquery to fetch the most recent confirmed reading prior to the current reading date
        prev_reading_sq = MeterReading.objects.filter(
            consumer=OuterRef('consumer'),
            is_confirmed=True,
            reading_date__lt=OuterRef('reading_date')
        ).order_by('-reading_date').values('reading_value')[:1]

        # Get all meter readings with related data AND annotated previous reading
        readings_queryset = MeterReading.objects.select_related(
            'consumer', 'consumer__barangay', 'submitted_by'
        ).annotate(
            prev_reading_value=Subquery(prev_reading_sq, output_field=FloatField())
        ).order_by('-reading_date', '-created_at')

        # Apply filters
        search_query = self.request.GET.get('search', '').strip()
        selected_barangay = self.request.GET.get('barangay', '')
        selected_status = self.request.GET.get('status', '')
        from_date = self.request.GET.get('from_date', '')
        to_date = self.request.GET.get('to_date', '')

        if search_query:
            readings_queryset = readings_queryset.filter(
                Q(consumer__first_name__icontains=search_query) |
                Q(consumer__last_name__icontains=search_query) |
                Q(consumer__id_number__icontains=search_query)
            )

        if selected_barangay:
            readings_queryset = readings_queryset.filter(consumer__barangay_id=selected_barangay)

        if selected_status:
            if selected_status == 'confirmed':
                readings_queryset = readings_queryset.filter(is_confirmed=True)
            elif selected_status == 'pending':
                readings_queryset = readings_queryset.filter(is_confirmed=False, is_rejected=False)

        if from_date:
            readings_queryset = readings_queryset.filter(reading_date__gte=from_date)

        if to_date:
            readings_queryset = readings_queryset.filter(reading_date__lte=to_date)

        # Prepare readings with consumption data efficiently without N+1 queries
        readings_with_data = []
        
        class MockPrevReading:
            def __init__(self, val):
                self.reading_value = val

        for reading in readings_queryset[:500]:  # Limit to recent 500 for performance
            if reading.prev_reading_value is None:
                baseline = reading.consumer.first_reading or 0
                prev_reading = MockPrevReading(baseline)
                consumption = reading.reading_value - baseline if reading.reading_value >= baseline else 0
            else:
                prev_reading = MockPrevReading(reading.prev_reading_value)
                consumption = reading.reading_value - reading.prev_reading_value

            readings_with_data.append({
                'reading': reading,
                'prev_reading': prev_reading,
                'consumption': consumption if reading.is_confirmed else (consumption if consumption >= 0 else 0),
                'display_id': reading.consumer.id_number
            })

        # Get pending readings for Pending tab, annotated with previous reading
        pending_readings = MeterReading.objects.filter(
            is_confirmed=False,
            is_rejected=False,
            source='app_manual'  # Manual entry from Smart Meter Reader app
        ).select_related('consumer', 'consumer__barangay', 'submitted_by').annotate(
            prev_reading_value=Subquery(prev_reading_sq, output_field=FloatField())
        ).order_by('-reading_date')

        # Calculate consumption directly from annotation
        for reading in pending_readings:
            if reading.prev_reading_value is not None:
                reading.previous_reading = reading.prev_reading_value
                reading.consumption = reading.reading_value - reading.prev_reading_value
            else:
                baseline = reading.consumer.first_reading or 0
                reading.previous_reading = baseline
                reading.consumption = reading.reading_value - baseline

        # Calculate statistics
        total_count = len(readings_with_data)
        confirmed_count = sum(1 for item in readings_with_data if item['reading'].is_confirmed)
        pending_count = sum(1 for item in readings_with_data if not item['reading'].is_confirmed and not item['reading'].is_rejected)

        # Calculate average consumption
        consumptions = [item['consumption'] for item in readings_with_data if item['consumption'] is not None and item['consumption'] > 0]
        avg_consumption = sum(consumptions) / len(consumptions) if consumptions else 0

        # Confirmed today count
        confirmed_today_count = MeterReading.objects.filter(
            is_confirmed=True,
            confirmed_at__date=today
        ).count()

        context.update({
            'readings': readings_with_data,  # Override the empty queryset
            'pending_readings': pending_readings,
            'barangays': barangays,
            'search_query': search_query,
            'selected_barangay': selected_barangay,
            'selected_status': selected_status,
            'from_date': from_date,
            'to_date': to_date,
            'total_count': total_count,
            'confirmed_count': confirmed_count,
            'pending_count': pending_count,
            'avg_consumption': avg_consumption,
            'confirmed_today_count': confirmed_today_count,
            'is_paginated': False,
        })
        
        return context

    def get(self, request, *args, **kwargs):
        # Check if export is requested
        if request.GET.get('export') == 'excel':
            return export_meter_readings_excel(request)
        return super().get(request, *args, **kwargs)



@login_required
def pending_readings_view(request):
    """
    Display all meter readings pending admin confirmation (with proof photos).
    """
    today = date.today()

    # --- FILTER PARAMS ---
    selected_barangay_id = request.GET.get('barangay', '')
    search_query = request.GET.get('search', '').strip()

    from django.db.models import OuterRef, Subquery, FloatField

    # Subquery to fetch the most recent confirmed reading
    prev_reading_sq = MeterReading.objects.filter(
        consumer=OuterRef('consumer'),
        is_confirmed=True,
        reading_date__lt=OuterRef('reading_date')
    ).order_by('-reading_date').values('reading_value')[:1]

    # Get pending readings using annotations to prevent N+1 queries
    pending_readings = MeterReading.objects.filter(
        is_confirmed=False,
        is_rejected=False,
        source='app_manual'  # Manual entry from Smart Meter Reader app
    ).select_related('consumer', 'consumer__barangay', 'submitted_by').annotate(
        prev_reading_value=Subquery(prev_reading_sq, output_field=FloatField())
    ).order_by('-reading_date')

    # Apply barangay filter
    if selected_barangay_id:
        pending_readings = pending_readings.filter(consumer__barangay_id=selected_barangay_id)

    # Apply name / ID search filter
    if search_query:
        pending_readings = pending_readings.filter(
            Q(consumer__first_name__icontains=search_query) |
            Q(consumer__last_name__icontains=search_query) |
            Q(consumer__id_number__icontains=search_query)
        )

    # Calculate consumption efficiently
    for reading in pending_readings:
        if reading.prev_reading_value is not None:
            reading.previous_reading = reading.prev_reading_value
            reading.consumption = reading.reading_value - reading.prev_reading_value
        else:
            baseline = reading.consumer.first_reading or 0
            reading.previous_reading = baseline
            reading.consumption = reading.reading_value - baseline

    # Stats
    confirmed_today_count = MeterReading.objects.filter(
        is_confirmed=True,
        confirmed_at__date=today
    ).count()

    rejected_today_count = MeterReading.objects.filter(
        is_rejected=True,
        rejected_at__date=today
    ).count()

    # All barangays for dropdown
    all_barangays = Barangay.objects.all().order_by('name')

    context = {
        'pending_readings': pending_readings,
        'confirmed_today_count': confirmed_today_count,
        'rejected_today_count': rejected_today_count,
        'all_barangays': all_barangays,
        'selected_barangay_id': selected_barangay_id,
        'search_query': search_query,
    }

    return render(request, 'consumers/pending_readings.html', context)


@login_required
def confirm_selected_readings(request, barangay_id):
    if request.method != "POST":
        return redirect('consumers:barangay_meter_readings', barangay_id=barangay_id)

    barangay = get_object_or_404(Barangay, id=barangay_id)
    reading_ids = request.POST.getlist('reading_ids')
    consumer_ids = request.POST.getlist('consumer_ids')  # For consumers with no reading (optional)

    # Debug: Check what readings were selected
    if not reading_ids:
        messages.warning(request, "No readings were selected for confirmation.")
        return redirect('consumers:barangay_meter_readings', barangay_id=barangay_id)

    success_count = 0

    # Process selected readings
    error_count = 0
    error_details = []
    for reading_id in reading_ids:
        try:
            reading = MeterReading.objects.get(id=reading_id)
            if reading.is_confirmed:
                continue

            consumer = reading.consumer
            prev = MeterReading.objects.filter(
                consumer=consumer,
                is_confirmed=True,
                reading_date__lt=reading.reading_date
            ).order_by('-reading_date').first()

            # Calculate consumption
            if prev:
                if reading.reading_value < prev.reading_value:
                    continue
                cons = reading.reading_value - prev.reading_value
            else:
                # First reading - use consumer's first_reading as baseline
                baseline = consumer.first_reading if consumer.first_reading else 0
                if reading.reading_value < baseline:
                    continue
                cons = reading.reading_value - baseline

            # Get system settings and calculate using TIERED RATES
            from ..utils import calculate_tiered_water_bill

            setting = SystemSetting.objects.first()
            if setting:
                billing_day = setting.billing_day_of_month
                due_day = setting.due_day_of_month
            else:
                billing_day = 1
                due_day = 20

            # Calculate bill using TIERED RATES
            total, average_rate, breakdown = calculate_tiered_water_bill(
                consumption=cons,
                usage_type=consumer.usage_type,
                settings=setting
            )

            Bill.objects.create(
                consumer=consumer,
                previous_reading=prev,
                current_reading=reading,
                billing_period=reading.reading_date.replace(day=billing_day),
                due_date=reading.reading_date.replace(day=due_day),
                consumption=cons,
                # Store ACTUAL tier breakdown
                tier1_consumption=breakdown['tier1_units'],
                tier1_amount=breakdown['tier1_amount'],
                tier2_consumption=breakdown['tier2_units'],
                tier2_rate=breakdown['tier2_rate'],
                tier2_amount=breakdown['tier2_amount'],
                tier3_consumption=breakdown['tier3_units'],
                tier3_rate=breakdown['tier3_rate'],
                tier3_amount=breakdown['tier3_amount'],
                tier4_consumption=breakdown['tier4_units'],
                tier4_rate=breakdown['tier4_rate'],
                tier4_amount=breakdown['tier4_amount'],
                tier5_consumption=breakdown['tier5_units'],
                tier5_rate=breakdown['tier5_rate'],
                tier5_amount=breakdown['tier5_amount'],
                rate_per_cubic=average_rate,
                fixed_charge=Decimal('0.00'),
                total_amount=total,
                status='Pending'
            )
            reading.is_confirmed = True
            reading.save()
            success_count += 1
        except Exception as e:
            import logging
            logging.error(f"Error confirming reading {reading_id}: {str(e)}")
            error_count += 1
            error_details.append(f"Reading {reading_id}: {str(e)}")
            continue

    if success_count > 0:
        messages.success(request, f"✅ {success_count} readings confirmed and bills generated.")
    elif error_count == 0 and success_count == 0:
        messages.info(request, "No new readings to confirm (may already be confirmed).")

    if error_count > 0:
        messages.error(request, f"⚠️ {error_count} readings failed: {'; '.join(error_details[:3])}")

    return redirect('consumers:barangay_meter_readings', barangay_id=barangay_id)
