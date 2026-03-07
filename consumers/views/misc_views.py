from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.conf import settings
from ..decorators import (
    get_client_ip, get_user_agent, is_admin_user, is_superuser_only,
    consumer_edit_permission_required, disconnect_permission_required,
    user_management_permission_required, system_settings_permission_required,
    billing_permission_required, reports_permission_required, view_only_for_admin,
    rate_limit_login, role_required
)
from django.db.models import Q, Max, Count, Sum, OuterRef, Subquery, Value, F
from django.db.models.functions import Concat, TruncMonth
from django.db import models
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone
from django.utils.decorators import method_decorator
from django.urls import reverse
from django.core.paginator import Paginator
from django.contrib.auth.models import User
from datetime import datetime, timedelta, date
try:
    from dateutil.relativedelta import relativedelta
except Exception:
    # Fallback: approximate relativedelta by using a timedelta of ~30 days per month
    # This keeps existing subtraction usages like relativedelta(months=5) working
    from datetime import timedelta as _td
    def relativedelta(months=0, **kwargs):
        return _td(days=30 * int(months))
from decimal import Decimal, InvalidOperation
import uuid
import json
import csv
import base64
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Cloudinary import with error handling (optional dependency)
try:
    from cloudinary import uploader as cloudinary_uploader  # type: ignore
    CLOUDINARY_AVAILABLE = True
except ImportError:
    cloudinary_uploader = None
    CLOUDINARY_AVAILABLE = False

from ..models import (
    Consumer, Barangay, Purok, MeterReading, Bill, SystemSetting, Payment,
    StaffProfile, UserLoginEvent, MeterBrand, PasswordResetToken, UserActivity,
    SystemSettingChangeLog, Notification
)
from ..forms import ConsumerForm


# Helper function to authenticate API requests using session token
def authenticate_api_request(request):
    """
    Authenticate API request using session token from Authorization header or request body.
    Returns the user if authenticated, None otherwise.
    """
    from django.contrib.sessions.models import Session

    token = None

    # Try Authorization header first (Bearer token)
    auth_header = request.META.get('HTTP_AUTHORIZATION', '')
    if auth_header.startswith('Bearer '):
        token = auth_header[7:]

    # Try request body if no header token
    if not token and request.body:
        try:
            data = json.loads(request.body.decode('utf-8'))
            token = data.get('token')
        except (json.JSONDecodeError, UnicodeDecodeError):
            pass

    if not token:
        return None

    try:
        # Find the session by key
        session = Session.objects.get(session_key=token)

        # Check if session is expired
        if session.expire_date < timezone.now():
            return None

        # Get user from session data
        session_data = session.get_decoded()
        user_id = session_data.get('_auth_user_id')

        if user_id:
            user = User.objects.get(id=user_id)
            return user
    except (Session.DoesNotExist, User.DoesNotExist):
        pass

    return None


# Helper function to get previous confirmed reading
def get_previous_reading(consumer):
    """Get the most recent confirmed meter reading for a consumer."""
    latest_reading = MeterReading.objects.filter(
        consumer=consumer,
        is_confirmed=True
    ).order_by('-reading_date', '-created_at').first()

    return latest_reading.reading_value if latest_reading else 0


# Helper function to calculate water bill
def calculate_water_bill(consumer, consumption):
    """
    Calculate water bill using TIERED rate structure from System Settings.

    Returns: (average_rate, total_amount, breakdown)
    - average_rate: Effective rate per cubic meter
    - total_amount: Total bill amount
    - breakdown: Dict with tier-by-tier calculation details
    """
    from ..utils import calculate_tiered_water_bill

    # Use tiered calculation from utils
    total_amount, average_rate, breakdown = calculate_tiered_water_bill(
        consumption=consumption,
        usage_type=consumer.usage_type
    )

    return float(average_rate), float(total_amount), breakdown



@csrf_exempt
@login_required
def api_create_reading(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
        consumer_id = data.get('consumer_id')
        reading_value = data.get('reading_value')
        reading_date = data.get('reading_date')  # YYYY-MM-DD

        # Validate consumer belongs to staff's barangay
        profile = StaffProfile.objects.get(user=request.user)
        consumer = Consumer.objects.get(id=consumer_id, barangay=profile.assigned_barangay)

        MeterReading.objects.create(
            consumer=consumer,
            reading_value=reading_value,
            reading_date=reading_date,
            source='field_app'
        )
        return JsonResponse({'status': 'success'})
    except StaffProfile.DoesNotExist:
        return JsonResponse({'error': 'Staff profile not found'}, status=403)
    except Consumer.DoesNotExist:
        return JsonResponse({'error': 'Consumer not found or not in assigned barangay'}, status=404)
    except Exception as e:
        # Log error but don't expose details
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"API create reading error: {e}", exc_info=True)
        return JsonResponse({'error': 'Failed to create reading'}, status=400)


# ... (your other views remain the same) ...
# NEW: View for field staff to see their assigned consumers
@login_required
def consumer_list_for_staff(request):
    """Display consumers for the logged-in staff member's assigned barangay."""
    try:
        profile = StaffProfile.objects.get(user=request.user)
        assigned_barangay = profile.assigned_barangay
        consumers = Consumer.objects.filter(barangay=assigned_barangay)
    except StaffProfile.DoesNotExist:
        messages.error(request, "User profile not found. Please contact an administrator.")
        return redirect('consumers:login') # Or another appropriate page

    # Get login time from session
    login_time_iso = request.session.get('login_time')
    login_time_str = None
    if login_time_iso:
        try:
            login_time_obj = timezone.datetime.fromisoformat(login_time_iso.replace('Z', '+00:00'))
            login_time_str = login_time_obj.strftime("%b %d, %Y %H:%M:%S")
        except ValueError:
            login_time_str = "Unknown"

    context = {
        'consumers': consumers,
        'assigned_barangay': assigned_barangay,
        'login_time': login_time_str,
        'user': request.user, # Pass user for username display
    }
    return render(request, 'consumers/consumer_list_for_staff.html', context) # Create this template


def user_logout(request):
    logout(request)
    messages.info(request, "You have been logged out.")
    return redirect('consumers:login') # Redirect to login page







def get_consumer_display_id(consumer):
    """Returns the consumer's ID number as the display ID"""
    return consumer.id_number or "—"



@login_required
def export_meter_readings_excel(request):
    """Export meter readings to Excel (.xlsx) file with logo and formatting"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    from datetime import datetime
    import os
    from django.conf import settings

    # Get filter parameters
    search_query = request.GET.get('search', '').strip()
    selected_barangay = request.GET.get('barangay', '')
    selected_status = request.GET.get('status', '')
    from_date = request.GET.get('from_date', '')
    to_date = request.GET.get('to_date', '')

    # Build queryset with same filters as the view
    readings_queryset = MeterReading.objects.select_related(
        'consumer', 'consumer__barangay', 'submitted_by'
    ).order_by('-reading_date', '-created_at')

    if search_query:
        readings_queryset = readings_queryset.filter(
            Q(consumer__first_name__icontains=search_query) |
            Q(consumer__last_name__icontains=search_query) |
            Q(consumer__id_number__icontains=search_query)
        )

    if selected_barangay:
        readings_queryset = readings_queryset.filter(consumer__barangay_id=selected_barangay)

    if selected_status:
        if selected_status == 'confirmed':
            readings_queryset = readings_queryset.filter(is_confirmed=True)
        elif selected_status == 'pending':
            readings_queryset = readings_queryset.filter(is_confirmed=False, is_rejected=False)

    if from_date:
        readings_queryset = readings_queryset.filter(reading_date__gte=from_date)

    if to_date:
        readings_queryset = readings_queryset.filter(reading_date__lte=to_date)

    # Limit to 2000 records for performance
    readings_queryset = readings_queryset[:2000]

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Meter Readings"

    # Try to add logo
    try:
        logo_path = os.path.join(settings.BASE_DIR, 'consumers', 'static', 'consumers', 'images', 'logo.png')
        if os.path.exists(logo_path):
            img = XLImage(logo_path)
            # Resize logo to fit in header (about 60x60 pixels)
            img.width = 60
            img.height = 60
            ws.add_image(img, 'A1')
    except Exception as e:
        pass  # If logo fails, continue without it

    # Styling
    title_font = Font(bold=True, size=16, color="003366")
    subtitle_font = Font(bold=True, size=11, color="0055aa")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Header Section (rows 1-5)
    ws.merge_cells('B1:G1')
    ws['B1'] = "BALILIHAN WATERWORKS"
    ws['B1'].font = title_font
    ws['B1'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('B2:G2')
    ws['B2'] = "METER READINGS REPORT"
    ws['B2'].font = subtitle_font
    ws['B2'].alignment = Alignment(horizontal='center', vertical='center')

    # Add filter info
    filter_info = []
    if selected_barangay:
        try:
            barangay = Barangay.objects.get(id=selected_barangay)
            filter_info.append(f"Barangay: {barangay.name}")
        except:
            pass
    if selected_status:
        filter_info.append(f"Status: {selected_status.title()}")
    if from_date and to_date:
        filter_info.append(f"Period: {from_date} to {to_date}")
    elif from_date:
        filter_info.append(f"From: {from_date}")
    elif to_date:
        filter_info.append(f"To: {to_date}")

    ws.merge_cells('B3:G3')
    ws['B3'] = " | ".join(filter_info) if filter_info else "All Records"
    ws['B3'].alignment = Alignment(horizontal='center')

    ws.merge_cells('B4:G4')
    ws['B4'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %I:%M %p')}"
    ws['B4'].alignment = Alignment(horizontal='center')
    ws['B4'].font = Font(size=9, color="666666")

    # Set row heights
    ws.row_dimensions[1].height = 50
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 18
    ws.row_dimensions[4].height = 18
    ws.row_dimensions[5].height = 5  # Empty row

    # Headers (row 6)
    headers = ['ID Number', 'Consumer Name', 'Barangay', 'Current', 'Previous', 'Consumption (m³)', 'Date', 'Source', 'Status']
    ws.append([])  # Row 5 (empty)
    ws.append(headers)  # Row 6

    header_row = ws[6]
    for cell in header_row:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # Data rows
    for reading in readings_queryset:
        # Get previous reading
        prev_reading = MeterReading.objects.filter(
            consumer=reading.consumer,
            is_confirmed=True,
            reading_date__lt=reading.reading_date
        ).order_by('-reading_date').first()

        # Calculate consumption
        if prev_reading:
            consumption = reading.reading_value - prev_reading.reading_value
            prev_value = prev_reading.reading_value
        elif reading.consumer.first_reading:
            consumption = reading.reading_value - reading.consumer.first_reading
            prev_value = reading.consumer.first_reading
        else:
            consumption = reading.reading_value
            prev_value = 0

        # Source display
        source_display = reading.get_source_display() if hasattr(reading, 'get_source_display') else reading.source

        # Status
        if reading.is_confirmed:
            status = "Confirmed"
        elif reading.is_rejected:
            status = "Rejected"
        else:
            status = "Pending"

        ws.append([
            reading.consumer.id_number or '—',
            f"{reading.consumer.first_name} {reading.consumer.last_name}",
            reading.consumer.barangay.name if reading.consumer.barangay else 'N/A',
            reading.reading_value,
            prev_value,
            consumption if reading.is_confirmed else (consumption if consumption >= 0 else 0),
            reading.reading_date.strftime('%Y-%m-%d'),
            source_display,
            status
        ])

    # Apply borders and alignment to all data cells
    for row in range(7, ws.max_row + 1):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row, col)
            cell.border = border
            if col in [4, 5, 6]:  # Numeric columns
                cell.alignment = Alignment(horizontal='center')
            else:
                cell.alignment = Alignment(horizontal='left')

    # Auto-adjust column widths
    column_widths = {
        'A': 15,  # ID Number
        'B': 30,  # Consumer Name
        'C': 20,  # Barangay
        'D': 12,  # Current
        'E': 12,  # Previous
        'F': 16,  # Consumption
        'G': 14,  # Date
        'H': 18,  # Source
        'I': 12,  # Status
    }
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    # Generate filename
    filename_parts = ['Meter_Readings']
    if selected_barangay:
        try:
            barangay = Barangay.objects.get(id=selected_barangay)
            filename_parts.append(barangay.name.replace(' ', '_'))
        except:
            pass
    if from_date and to_date:
        filename_parts.append(f"{from_date}_to_{to_date}")
    filename_parts.append(datetime.now().strftime('%Y%m%d'))

    filename = '_'.join(filename_parts) + '.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response



@login_required
@user_management_permission_required
def archived_users(request):
    """
    View list of archived/deleted users.
    RESTRICTED: Superuser only.
    """
    from ..models import ArchivedUser

    # Get search query parameter
    search_query = request.GET.get('search', '').strip()

    # Get all archived users
    archived_list = ArchivedUser.objects.all().order_by('-archived_at')

    # Apply search filter if provided
    if search_query:
        archived_list = archived_list.filter(
            Q(username__icontains=search_query) |
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(email__icontains=search_query)
        )

    # Pagination
    paginator = Paginator(archived_list, 20)
    page = request.GET.get('page', 1)
    try:
        archived_users_page = paginator.page(page)
    except:
        archived_users_page = paginator.page(1)

    context = {
        'archived_users': archived_users_page,
        'search_query': search_query,
        'total_archived': ArchivedUser.objects.count(),
    }

    return render(request, 'consumers/archived_users.html', context)



@login_required
@user_management_permission_required
def permanently_delete_archived_user(request, archived_id):
    """
    Permanently delete an archived user record.
    RESTRICTED: Superuser only.
    """
    from ..models import ArchivedUser

    archived_user = get_object_or_404(ArchivedUser, id=archived_id)

    if request.method == 'POST':
        username = archived_user.username
        archived_user.delete()
        messages.success(request, f"Archived user '{username}' has been permanently deleted.")
        return redirect('consumers:archived_users')

    return redirect('consumers:archived_users')
