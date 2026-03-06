from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.conf import settings
from ..decorators import (
    get_client_ip, get_user_agent, is_admin_user, is_superuser_only,
    consumer_edit_permission_required, disconnect_permission_required,
    user_management_permission_required, system_settings_permission_required,
    billing_permission_required, reports_permission_required, view_only_for_admin,
    rate_limit_login, role_required
)
from django.db.models import Q, Max, Count, Sum, OuterRef, Subquery, Value, F
from django.db.models.functions import Concat, TruncMonth
from django.db import models
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone
from django.utils.decorators import method_decorator
from django.urls import reverse
from django.core.paginator import Paginator
from django.contrib.auth.models import User
from datetime import datetime, timedelta, date
try:
    from dateutil.relativedelta import relativedelta
except Exception:
    # Fallback: approximate relativedelta by using a timedelta of ~30 days per month
    # This keeps existing subtraction usages like relativedelta(months=5) working
    from datetime import timedelta as _td
    def relativedelta(months=0, **kwargs):
        return _td(days=30 * int(months))
from decimal import Decimal, InvalidOperation
import uuid
import json
import csv
import base64
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Cloudinary import with error handling (optional dependency)
try:
    from cloudinary import uploader as cloudinary_uploader  # type: ignore
    CLOUDINARY_AVAILABLE = True
except ImportError:
    cloudinary_uploader = None
    CLOUDINARY_AVAILABLE = False

from ..models import (
    Consumer, Barangay, Purok, MeterReading, Bill, SystemSetting, Payment,
    StaffProfile, UserLoginEvent, MeterBrand, PasswordResetToken, UserActivity,
    SystemSettingChangeLog, Notification
)
from ..forms import ConsumerForm


# Helper function to authenticate API requests using session token
def authenticate_api_request(request):
    """
    Authenticate API request using session token from Authorization header or request body.
    Returns the user if authenticated, None otherwise.
    """
    from django.contrib.sessions.models import Session

    token = None

    # Try Authorization header first (Bearer token)
    auth_header = request.META.get('HTTP_AUTHORIZATION', '')
    if auth_header.startswith('Bearer '):
        token = auth_header[7:]

    # Try request body if no header token
    if not token and request.body:
        try:
            data = json.loads(request.body.decode('utf-8'))
            token = data.get('token')
        except (json.JSONDecodeError, UnicodeDecodeError):
            pass

    if not token:
        return None

    try:
        # Find the session by key
        session = Session.objects.get(session_key=token)

        # Check if session is expired
        if session.expire_date < timezone.now():
            return None

        # Get user from session data
        session_data = session.get_decoded()
        user_id = session_data.get('_auth_user_id')

        if user_id:
            user = User.objects.get(id=user_id)
            return user
    except (Session.DoesNotExist, User.DoesNotExist):
        pass

    return None


# Helper function to get previous confirmed reading
def get_previous_reading(consumer):
    """Get the most recent confirmed meter reading for a consumer."""
    latest_reading = MeterReading.objects.filter(
        consumer=consumer,
        is_confirmed=True
    ).order_by('-reading_date', '-created_at').first()

    return latest_reading.reading_value if latest_reading else 0


# Helper function to calculate water bill
def calculate_water_bill(consumer, consumption):
    """
    Calculate water bill using TIERED rate structure from System Settings.

    Returns: (average_rate, total_amount, breakdown)
    - average_rate: Effective rate per cubic meter
    - total_amount: Total bill amount
    - breakdown: Dict with tier-by-tier calculation details
    """
    from .utils import calculate_tiered_water_bill

    # Use tiered calculation from utils
    total_amount, average_rate, breakdown = calculate_tiered_water_bill(
        consumption=consumption,
        usage_type=consumer.usage_type
    )

    return float(average_rate), float(total_amount), breakdown



@login_required
def reports(request):
    """
    Income dashboard showing monthly and yearly income — all barangays combined and per barangay.
    """
    import calendar as cal
    from django.db.models.functions import ExtractMonth

    now = datetime.now()
    current_year = now.year

    # --- Filters: year, month_from, month_to ---
    year = request.GET.get('year', current_year)
    try:
        year = int(year)
    except (ValueError, TypeError):
        year = current_year

    month_from = request.GET.get('month_from', 1)
    month_to = request.GET.get('month_to', 12)
    try:
        month_from = int(month_from)
        month_to = int(month_to)
    except (ValueError, TypeError):
        month_from, month_to = 1, 12
    month_from = max(1, min(12, month_from))
    month_to = max(month_from, min(12, month_to))

    # Year choices
    earliest_payment = Payment.objects.order_by('payment_date').first()
    start_year = earliest_payment.payment_date.year if earliest_payment else current_year
    year_choices = list(range(current_year, start_year - 1, -1))

    # Month choices
    month_choices = [{'num': m, 'name': cal.month_name[m]} for m in range(1, 13)]
    month_names = [cal.month_name[m] for m in range(1, 13)]

    # --- All Barangays: monthly totals (income + consumption) ---
    monthly_qs = (
        Payment.objects.filter(
            payment_date__year=year,
            payment_date__month__gte=month_from,
            payment_date__month__lte=month_to,
        )
        .annotate(month=ExtractMonth('payment_date'))
        .values('month')
        .annotate(total=Sum('amount_paid'), consumption=Sum('bill__consumption'))
        .order_by('month')
    )
    monthly_dict = {row['month']: row for row in monthly_qs}
    monthly_data = []
    range_total_income = 0
    range_total_consumption = 0
    for m in range(month_from, month_to + 1):
        row = monthly_dict.get(m, {})
        amount = row.get('total', 0) or 0
        cons = row.get('consumption', 0) or 0
        monthly_data.append({'month': month_names[m - 1], 'total': amount, 'consumption': cons})
        range_total_income += amount
        range_total_consumption += cons

    # Range display label
    if month_from == month_to:
        range_label = f"{cal.month_name[month_from]} {year}"
    else:
        range_label = f"{cal.month_abbr[month_from]} – {cal.month_abbr[month_to]} {year}"

    # --- Per Barangay: monthly totals (income + consumption) ---
    per_brgy_qs = (
        Payment.objects.filter(
            payment_date__year=year,
            payment_date__month__gte=month_from,
            payment_date__month__lte=month_to,
        )
        .annotate(month=ExtractMonth('payment_date'))
        .values('bill__consumer__barangay__id', 'bill__consumer__barangay__name', 'month')
        .annotate(total=Sum('amount_paid'), consumption=Sum('bill__consumption'))
        .order_by('bill__consumer__barangay__name', 'month')
    )
    brgy_map = {}
    for row in per_brgy_qs:
        brgy_id = row['bill__consumer__barangay__id']
        brgy_name = row['bill__consumer__barangay__name']
        if brgy_id not in brgy_map:
            brgy_map[brgy_id] = {'name': brgy_name, 'months': {}, 'range_total': 0, 'range_consumption': 0}
        amount = row['total'] or 0
        cons = row['consumption'] or 0
        brgy_map[brgy_id]['months'][row['month']] = {'total': amount, 'consumption': cons}
        brgy_map[brgy_id]['range_total'] += amount
        brgy_map[brgy_id]['range_consumption'] += cons

    barangay_data = []
    for brgy_id in sorted(brgy_map, key=lambda x: brgy_map[x]['name']):
        entry = brgy_map[brgy_id]
        months = []
        for m in range(month_from, month_to + 1):
            md = entry['months'].get(m, {})
            months.append({
                'month': month_names[m - 1],
                'total': md.get('total', 0) or 0,
                'consumption': md.get('consumption', 0) or 0,
            })
        barangay_data.append({
            'name': entry['name'],
            'range_total': entry['range_total'],
            'range_consumption': entry['range_consumption'],
            'months': months,
        })

    barangays = Barangay.objects.all().order_by('name')

    context = {
        'year': year,
        'year_choices': year_choices,
        'month_choices': month_choices,
        'month_from': month_from,
        'month_to': month_to,
        'range_label': range_label,
        'monthly_data': monthly_data,
        'range_total_income': range_total_income,
        'range_total_consumption': range_total_consumption,
        'barangay_data': barangay_data,
        'barangays': barangays,
    }

    return render(request, 'consumers/reports.html', context)



@login_required
def barangay_report(request, barangay_id):
    """
    Ledger-style barangay report showing 12-month billing tables per consumer.
    Replicates the physical ledger book used by the waterworks office.
    """
    from datetime import date
    import calendar

    barangay = get_object_or_404(Barangay, id=barangay_id)

    # Year selector (default: current year)
    year = request.GET.get('year')
    try:
        year = int(year)
    except (TypeError, ValueError):
        year = date.today().year

    # Get all active consumers in this barangay, sorted by last name
    consumers = Consumer.objects.filter(
        barangay=barangay,
        status='active'
    ).order_by('last_name', 'first_name')

    # Build ledger data for each consumer
    month_names = [calendar.month_name[m] for m in range(1, 13)]
    consumer_ledger = []

    for consumer in consumers:
        # Get all bills for this consumer in the selected year
        bills = Bill.objects.filter(
            consumer=consumer,
            billing_period__year=year
        ).select_related('consumer', 'current_reading')

        # Get all payments for bills in this year
        payments = Payment.objects.filter(
            bill__consumer=consumer,
            bill__billing_period__year=year
        ).select_related('bill')

        # Build month-by-month data
        months_data = []
        for month_num in range(1, 13):
            bill = bills.filter(billing_period__month=month_num).first()
            payment = payments.filter(bill__billing_period__month=month_num).first()

            month_entry = {
                'month': month_names[month_num - 1],
                'consumption': bill.consumption if bill else '',
                'amount_due': bill.total_amount if bill else '',
                'penalty': bill.effective_penalty if bill and bill.effective_penalty > 0 else '',
                'amount_paid': payment.amount_paid if payment else '',
                'receipt_number': payment.or_number if payment else '',
                'date_issued': payment.payment_date if payment else '',
                'reading': bill.current_reading.reading_value if bill and bill.current_reading else '',
                'initial': '',
            }
            months_data.append(month_entry)

        consumer_ledger.append({
            'consumer': consumer,
            'months': months_data,
        })

    # Year range for selector (registration year of earliest consumer to current year + 1)
    current_year = date.today().year
    earliest = Consumer.objects.filter(barangay=barangay).order_by('registration_date').first()
    start_year = earliest.registration_date.year if earliest else current_year
    year_choices = list(range(current_year + 1, start_year - 1, -1))

    context = {
        'barangay': barangay,
        'year': year,
        'year_choices': year_choices,
        'consumer_ledger': consumer_ledger,
        'consumer_count': consumers.count(),
    }

    return render(request, 'consumers/barangay_report.html', context)



@login_required
def export_report_excel(request):
    """Export report as Excel (.xlsx) file with formatting"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    from datetime import datetime

    report_type = request.GET.get('report_type', 'revenue')
    date_from_str = request.GET.get('date_from')
    date_to_str = request.GET.get('date_to')

    if not date_from_str or not date_to_str:
        return HttpResponse("Date range parameters required", status=400)

    try:
        date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
        date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date()
    except:
        return HttpResponse("Invalid date format", status=400)

    # Create workbook
    wb = Workbook()
    ws = wb.active

    # Date range for display
    date_range_display = f"{date_from.strftime('%B %d, %Y')} - {date_to.strftime('%B %d, %Y')}"
    date_range_short = f"{date_from.strftime('%Y%m%d')}-{date_to.strftime('%Y%m%d')}"

    # Styling
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    title_font = Font(bold=True, size=14)
    total_font = Font(bold=True, size=11)
    total_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    if report_type == 'revenue':
        # Revenue Report
        ws.title = f"Revenue Report"

        # Title
        ws['A1'] = "BALILIHAN WATERWORKS - REVENUE REPORT"
        ws['A1'].font = title_font
        ws['A2'] = f"Period: {date_range_display}"
        ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %I:%M %p')}"

        # Headers
        headers = ['OR Number', 'Consumer Name', 'ID Number', 'Payment Date', 'Amount Paid', 'Change Given', 'Total Received']
        ws.append([])  # Empty row
        ws.append(headers)

        header_row = ws[5]
        for cell in header_row:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # Data
        payments = Payment.objects.filter(
            payment_date__gte=date_from,
            payment_date__lte=date_to
        ).select_related('bill__consumer').order_by('payment_date')

        total_amount = 0
        total_change = 0
        total_received = 0

        for payment in payments:
            consumer = payment.bill.consumer
            ws.append([
                payment.or_number,
                consumer.full_name,
                consumer.id_number or '—',
                payment.payment_date.strftime('%Y-%m-%d'),
                float(payment.amount_paid),
                float(payment.change),
                float(payment.received_amount)
            ])
            total_amount += payment.amount_paid
            total_change += payment.change
            total_received += payment.received_amount

        # Total row
        total_row = ws.max_row + 1
        ws[f'A{total_row}'] = 'TOTAL'
        ws[f'A{total_row}'].font = total_font
        ws[f'E{total_row}'] = float(total_amount)
        ws[f'F{total_row}'] = float(total_change)
        ws[f'G{total_row}'] = float(total_received)

        for col in ['A', 'E', 'F', 'G']:
            ws[f'{col}{total_row}'].font = total_font
            ws[f'{col}{total_row}'].fill = total_fill
            ws[f'{col}{total_row}'].border = border

        # Format currency columns
        for row in range(6, ws.max_row + 1):
            for col in ['E', 'F', 'G']:
                ws[f'{col}{row}'].number_format = '₱#,##0.00'
                ws[f'{col}{row}'].border = border
                ws[f'{col}{row}'].alignment = Alignment(horizontal='right')

        filename = f"Revenue_Report_{date_range_short}.xlsx"

    elif report_type == 'delinquency':
        # Delinquency Report
        ws.title = f"Delinquency Report"

        # Title
        ws['A1'] = "BALILIHAN WATERWORKS - DELINQUENT ACCOUNTS REPORT"
        ws['A1'].font = title_font
        ws['A2'] = f"Period: {date_range_display}"
        ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %I:%M %p')}"

        # Headers
        headers = ['ID Number', 'Consumer Name', 'Barangay', 'Billing Period', 'Due Date', 'Amount Due', 'Status']
        ws.append([])
        ws.append(headers)

        header_row = ws[5]
        for cell in header_row:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # Data
        bills = Bill.objects.filter(
            billing_period__gte=date_from,
            billing_period__lte=date_to,
            status__in=['Pending', 'Overdue']
        ).select_related('consumer__barangay').order_by('consumer__id_number')

        total_due = 0

        for bill in bills:
            consumer = bill.consumer
            ws.append([
                consumer.id_number or '—',
                consumer.full_name,
                consumer.barangay.name if consumer.barangay else 'N/A',
                bill.billing_period.strftime('%B %Y'),
                bill.due_date.strftime('%Y-%m-%d'),
                float(bill.total_amount),
                bill.status
            ])
            total_due += bill.total_amount

        # Total row
        total_row = ws.max_row + 1
        ws[f'A{total_row}'] = 'TOTAL DELINQUENT AMOUNT'
        ws[f'A{total_row}'].font = total_font
        ws[f'F{total_row}'] = float(total_due)
        ws[f'F{total_row}'].font = total_font
        ws[f'F{total_row}'].fill = total_fill
        ws[f'F{total_row}'].border = border
        ws[f'F{total_row}'].number_format = '₱#,##0.00'

        # Format currency column
        for row in range(6, ws.max_row):
            ws[f'F{row}'].number_format = '₱#,##0.00'
            ws[f'F{row}'].border = border
            ws[f'F{row}'].alignment = Alignment(horizontal='right')
            for col in ['A', 'B', 'C', 'D', 'E', 'G']:
                ws[f'{col}{row}'].border = border

        filename = f"Delinquency_Report_{date_range_short}.xlsx"

    elif report_type == 'summary':
        # Summary Report
        ws.title = f"Summary Report"

        # Title
        ws['A1'] = "BALILIHAN WATERWORKS - PAYMENT SUMMARY REPORT"
        ws['A1'].font = title_font
        ws['A2'] = f"Period: {date_range_display}"
        ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %I:%M %p')}"

        # Headers
        headers = ['ID Number', 'Consumer Name', 'Total Amount Paid', 'Number of Payments']
        ws.append([])
        ws.append(headers)

        header_row = ws[5]
        for cell in header_row:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # Data
        summary_data = Payment.objects.filter(
            payment_date__gte=date_from,
            payment_date__lte=date_to
        ).values('bill__consumer__id_number').annotate(
            bill__consumer__full_name=Concat(
                'bill__consumer__first_name',
                Value(' '),
                'bill__consumer__middle_name',
                Value(' '),
                'bill__consumer__last_name'
            ),
            total_paid=Sum('amount_paid'),
            count=Count('id')
        ).order_by('bill__consumer__id_number')

        total_amount = 0
        total_count = 0

        for item in summary_data:
            ws.append([
                item['bill__consumer__id_number'] or '—',
                item['bill__consumer__full_name'],
                float(item['total_paid']),
                item['count']
            ])
            total_amount += item['total_paid']
            total_count += item['count']

        # Total row
        total_row = ws.max_row + 1
        ws[f'A{total_row}'] = 'TOTAL'
        ws[f'A{total_row}'].font = total_font
        ws[f'C{total_row}'] = float(total_amount)
        ws[f'D{total_row}'] = total_count

        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{total_row}'].font = total_font
            ws[f'{col}{total_row}'].fill = total_fill
            ws[f'{col}{total_row}'].border = border

        # Format currency column
        for row in range(6, ws.max_row + 1):
            ws[f'C{row}'].number_format = '₱#,##0.00'
            ws[f'C{row}'].alignment = Alignment(horizontal='right')
            for col in ['A', 'B', 'C', 'D']:
                ws[f'{col}{row}'].border = border

        filename = f"Payment_Summary_{date_range_short}.xlsx"

    else:
        return HttpResponse("Invalid report type", status=400)

    # Auto-size columns
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save to response
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response
